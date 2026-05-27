# -*- coding: utf-8 -*-
"""Tests for drawing batch scan/match orchestration."""

from datetime import datetime
from pathlib import Path

import pytest

from src.services.comparison.base import ChangeRecord, ChangeType, ComparisonResult
from src.services.comparison.drawing_batch import (
    BatchCompareItemResult,
    BatchCompareJob,
    BatchCompareOptions,
    BatchCompareSummary,
    DescriptorBuildOptions,
    DrawingFileDescriptor,
    DrawingKind,
    FilenameIdentity,
    MatchCandidate,
    MatchStatus,
    MatchingOptions,
    apply_manual_matches,
    compare_pdf_documents,
    compare_candidate,
    confirmed_pair_uniqueness_violations,
    match_drawing_sets,
    parse_filename_identity,
    scan_drawing_inputs,
    score_match,
    write_manual_match_csv,
    _legacy_ezdxf_fallback_available,
)
from src.services.comparison.pair_identity import candidate_pair_uuid


def _descriptor(
    name: str,
    *,
    kind: DrawingKind = DrawingKind.CAD,
    drawing_number: str | None = None,
    sheet: str | None = None,
    text_hints: tuple[str, ...] = tuple(),
    layers: tuple[str, ...] = tuple(),
    entity_counts: dict[str, int] | None = None,
    visual_fingerprint: str = "",
    page_count: int = 0,
) -> DrawingFileDescriptor:
    identity = parse_filename_identity(name)
    if drawing_number or sheet:
        identity = FilenameIdentity(
            original_stem=identity.original_stem,
            match_key=identity.match_key,
            tokens=identity.tokens,
            revision=identity.revision,
            drawing_number=drawing_number or identity.drawing_number,
            sheet=sheet or identity.sheet,
        )
    return DrawingFileDescriptor(
        path=str(Path("C:/drawings") / name),
        kind=kind,
        extension=Path(name).suffix.lower(),
        relative_path=name,
        identity=identity,
        text_hints=text_hints,
        layers=layers,
        entity_counts=entity_counts or {},
        visual_fingerprint=visual_fingerprint,
        page_count=page_count,
    )


def _confirmed_candidate(name: str) -> MatchCandidate:
    return MatchCandidate(
        source_a=_descriptor(f"{name}.dwg"),
        source_b=_descriptor(f"{name}_REV1.dwg"),
        score=0.9,
        status=MatchStatus.AUTO_CONFIRMED,
    )


def test_revision_only_difference_auto_matches() -> None:
    old = _descriptor("S-101_REV60.dwg")
    new = _descriptor("S-101_REV66.dwg")

    matches = match_drawing_sets([old], [new])

    assert len(matches) == 1
    assert matches[0].status == MatchStatus.AUTO_CONFIRMED
    assert matches[0].score >= 0.85
    assert old.identity.match_key == new.identity.match_key


def test_project_drawing_number_is_extracted_before_prefix() -> None:
    assert parse_filename_identity("S21-0031_P STRUCTURAL PLAN 3F.dwg").drawing_number == "S21-0031"
    assert parse_filename_identity("S20-0001_P TEST.dwg").drawing_number == "S20-0001"
    assert parse_filename_identity("S-101_REV60.dwg").drawing_number == "S101"


def test_full_project_code_exact_match_wins_over_similar_prefix_candidates() -> None:
    source_a = _descriptor("S21-0031_P JISANG3 STRUCTURAL PLAN-1.dwg")
    correct_b = _descriptor("S21-0031_P 3F STRUCTURAL PLAN(1).dwg")
    wrong_b = _descriptor("S21-0102_P 10F STRUCTURAL PLAN(2).dwg")

    matches = match_drawing_sets([source_a], [wrong_b, correct_b])
    paired = [match for match in matches if match.source_a and match.source_b]

    assert len(paired) == 1
    assert paired[0].source_b == correct_b
    assert paired[0].score >= 0.85
    assert not any(
        match.source_a == source_a and match.source_b == wrong_b
        for match in matches
    )


def test_full_project_code_mismatch_does_not_create_review_candidate() -> None:
    source_a = _descriptor(
        "S21-0121_P 12F STRUCTURAL PLAN-1.dwg",
        text_hints=("S21-0121 STRUCTURAL PLAN",),
        layers=("GRID", "BEAM"),
        entity_counts={"LINE": 20, "TEXT": 2},
    )
    wrong_b = _descriptor(
        "S21-0031_P 3F STRUCTURAL PLAN-1.dwg",
        text_hints=("S21-0031 STRUCTURAL PLAN",),
        layers=("GRID", "BEAM"),
        entity_counts={"LINE": 20, "TEXT": 2},
    )

    scored = score_match(source_a, wrong_b)
    matches = match_drawing_sets([source_a], [wrong_b])

    assert scored.score < MatchingOptions().review_threshold
    assert any("drawing code mismatch" in reason for reason in scored.reasons)
    assert {match.status for match in matches} == {
        MatchStatus.UNMATCHED_A,
        MatchStatus.UNMATCHED_B,
    }


def test_same_drawing_number_from_metadata_auto_matches() -> None:
    old = _descriptor(
        "old_plan.dxf",
        drawing_number="S101",
        sheet="B2F",
        text_hints=("DRAWING NO S-101 STRUCTURAL PLAN B2F",),
        layers=("GRID", "BEAM"),
        entity_counts={"LINE": 20, "TEXT": 2},
    )
    new = _descriptor(
        "new_structure.dwg",
        drawing_number="S101",
        sheet="B2F",
        text_hints=("DRAWING NO S-101 STRUCTURAL PLAN B2F",),
        layers=("GRID", "BEAM"),
        entity_counts={"LINE": 20, "TEXT": 2},
    )

    matches = match_drawing_sets([old], [new])

    assert matches[0].status == MatchStatus.AUTO_CONFIRMED
    assert any("drawing number matched" in reason for reason in matches[0].reasons)


def test_pdf_text_and_drawing_number_can_match_different_filenames() -> None:
    old = _descriptor(
        "submission_a.pdf",
        kind=DrawingKind.PDF,
        drawing_number="A200",
        sheet="1F",
        text_hints=("DRAWING NO A-200 ARCHITECTURAL 1F PLAN",),
        visual_fingerprint="ff00ff00ff00ff00",
        page_count=2,
    )
    new = _descriptor(
        "permit_set_sheet_01.pdf",
        kind=DrawingKind.PDF,
        drawing_number="A200",
        sheet="1F",
        text_hints=("DRAWING NO A-200 ARCHITECTURAL 1F PLAN",),
        visual_fingerprint="ff00ff00ff00ff00",
        page_count=2,
    )

    matches = match_drawing_sets([old], [new])

    assert matches[0].status == MatchStatus.AUTO_CONFIRMED
    assert matches[0].score >= 0.85


def test_pdf_compare_records_page_bbox_metadata(tmp_path: Path) -> None:
    fitz = pytest.importorskip("fitz")
    old_pdf = tmp_path / "old.pdf"
    new_pdf = tmp_path / "new.pdf"
    doc = fitz.open()
    doc.new_page(width=200, height=100)
    doc.save(old_pdf)
    doc.close()
    doc = fitz.open()
    doc.new_page(width=200, height=100)
    doc.new_page(width=200, height=100)
    doc.save(new_pdf)
    doc.close()

    result = compare_pdf_documents(
        old_pdf,
        new_pdf,
        BatchCompareOptions(compare_pdf_all_pages=False, pdf_text_compare=False, pdf_dpi=72),
    )

    added_page = next(change for change in result.changes if change.key == "page_1_added")
    assert added_page.metadata["source_format"] == "pdf"
    assert added_page.metadata["entity_type"] == "PDF_PAGE"
    assert added_page.metadata["pdf_page"] == 1
    assert added_page.metadata["pdf_dpi"] == 72
    assert added_page.metadata["bbox_coordinate_space"] == "image_pixels"
    assert added_page.metadata["bbox"] == [0.0, 0.0, 200.0, 100.0]


def test_pdf_compare_does_not_probe_ocr_when_fallback_disabled(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    fitz = pytest.importorskip("fitz")
    old_pdf = tmp_path / "old.pdf"
    new_pdf = tmp_path / "new.pdf"
    for path in (old_pdf, new_pdf):
        doc = fitz.open()
        doc.new_page(width=200, height=100)
        doc.save(path)
        doc.close()

    def fail_ocr_probe() -> bool:
        raise AssertionError("OCR probe should be opt-in for PDF batch compare")

    monkeypatch.setattr(
        "src.services.comparison.drawing_differ._check_ocr_lazily",
        fail_ocr_probe,
    )

    result = compare_pdf_documents(
        old_pdf,
        new_pdf,
        BatchCompareOptions(
            compare_pdf_all_pages=False,
            pdf_text_compare=True,
            pdf_dpi=72,
            use_ocr_fallback=False,
        ),
    )

    assert result.metadata["ocr_fallback_enabled"] is False


def test_ambiguous_candidates_require_review() -> None:
    source_a = _descriptor("S-101_REV0.dwg")
    first_b = _descriptor("S-101_REV1.dwg")
    second_b = _descriptor("S-101_COPY_REV2.dwg")

    matches = match_drawing_sets([source_a], [first_b, second_b])
    paired = [match for match in matches if match.source_a and match.source_b]

    assert len(paired) == 1
    assert paired[0].status == MatchStatus.REVIEW_REQUIRED
    assert paired[0].alternates
    assert paired[0].alternates[0].source_b.name in {"S-101_REV1.dwg", "S-101_COPY_REV2.dwg"}
    assert any("ambiguous competing candidate" in reason for reason in paired[0].reasons)


def test_unique_exact_drawing_number_auto_confirms_despite_similar_pdf_alternates() -> None:
    source_a = _descriptor(
        "S-2401_REV0.pdf",
        kind=DrawingKind.PDF,
        drawing_number="S2401",
        sheet="REV0",
        text_hints=("SLAB PLAN GRID A B D13@100",),
        visual_fingerprint="ff00ff00ff00ff00",
        page_count=1,
    )
    correct_b = _descriptor(
        "S-2401_REV1.pdf",
        kind=DrawingKind.PDF,
        drawing_number="S2401",
        sheet="REV1",
        text_hints=("SLAB PLAN GRID A B D13@200",),
        visual_fingerprint="ff00ff00ff00ff01",
        page_count=1,
    )
    similar_b = _descriptor(
        "S-2407_REV1.pdf",
        kind=DrawingKind.PDF,
        drawing_number="S2407",
        sheet="REV1",
        text_hints=("SLAB PLAN GRID A B D13@200",),
        visual_fingerprint="ff00ff00ff00ff01",
        page_count=1,
    )

    matches = match_drawing_sets([source_a], [similar_b, correct_b])
    paired = [match for match in matches if match.source_a and match.source_b]

    assert len(paired) == 1
    assert paired[0].source_b == correct_b
    assert paired[0].status == MatchStatus.AUTO_CONFIRMED
    assert paired[0].alternates
    assert not any("ambiguous competing candidate" in reason for reason in paired[0].reasons)


def test_alternate_limit_can_be_configured() -> None:
    source_a = _descriptor("S-101_REV0.dwg")
    candidates_b = [
        _descriptor("S-101_REV1.dwg"),
        _descriptor("S-101_COPY_REV2.dwg"),
        _descriptor("S-101_FINAL_REV3.dwg"),
    ]

    matches = match_drawing_sets(
        [source_a],
        candidates_b,
        options=MatchingOptions(alternate_limit=1),
    )
    paired = [match for match in matches if match.source_a and match.source_b]

    assert len(paired[0].alternates) == 1


def test_incompatible_cad_pdf_pair_is_not_created() -> None:
    source_a = _descriptor("S-101.dwg", kind=DrawingKind.CAD)
    source_b = _descriptor("S-101.pdf", kind=DrawingKind.PDF)

    matches = match_drawing_sets([source_a], [source_b])

    assert {match.status for match in matches} == {
        MatchStatus.UNMATCHED_A,
        MatchStatus.UNMATCHED_B,
    }
    assert not any(match.source_a and match.source_b for match in matches)


def test_scan_folder_can_include_or_exclude_subfolders(tmp_path) -> None:
    root = tmp_path / "drawings"
    nested = root / "nested"
    nested.mkdir(parents=True)
    (root / "A-100.dxf").write_text("invalid dxf placeholder", encoding="utf-8")
    (nested / "A-101.dxf").write_text("invalid dxf placeholder", encoding="utf-8")
    (root / "notes.txt").write_text("ignored", encoding="utf-8")

    top_level = scan_drawing_inputs(root, DescriptorBuildOptions(recursive=False))
    recursive = scan_drawing_inputs(root, DescriptorBuildOptions(recursive=True))

    assert [descriptor.name for descriptor in top_level] == ["A-100.dxf"]
    assert [descriptor.name for descriptor in recursive] == ["A-100.dxf", "A-101.dxf"]


def test_batch_job_runs_confirmed_pairs(monkeypatch) -> None:
    source_a = _descriptor("S-101.dwg")
    source_b = _descriptor("S-101_REV1.dwg")
    candidate = MatchCandidate(
        source_a=source_a,
        source_b=source_b,
        score=0.9,
        status=MatchStatus.AUTO_CONFIRMED,
    )

    def fake_compare(candidate, options):
        result = ComparisonResult(source_a=candidate.source_a.path, source_b=candidate.source_b.path)
        result.add_change(ChangeRecord(key="line_1", change_type=ChangeType.MODIFIED))
        return result

    monkeypatch.setattr(
        "src.services.comparison.drawing_batch.compare_candidate",
        fake_compare,
    )

    summary = BatchCompareJob([candidate]).run()

    assert summary.completed_pairs == 1
    assert summary.failed_pairs == 0
    assert summary.total_changes == 1


def test_batch_job_marks_failed_comparison_result_as_failed(monkeypatch) -> None:
    candidate = _confirmed_candidate("S-101")

    def fake_compare(candidate, options):
        result = ComparisonResult(source_a=candidate.source_a.path, source_b=candidate.source_b.path)
        result.metadata.update(
            {
                "pipeline_status": "failed",
                "error_code": "COMPARE_IMPORT_FAILED",
                "message": "CAD compare import failed",
            }
        )
        return result

    monkeypatch.setattr(
        "src.services.comparison.drawing_batch.compare_candidate",
        fake_compare,
    )

    summary = BatchCompareJob([candidate]).run()

    assert summary.completed_pairs == 0
    assert summary.failed_pairs == 1
    assert summary.items[0].error == "COMPARE_IMPORT_FAILED: CAD compare import failed"


def test_batch_job_can_skip_full_compare_state_json(tmp_path, monkeypatch) -> None:
    candidate = _confirmed_candidate("S-101")

    def fake_compare(candidate, options):
        result = ComparisonResult(source_a=candidate.source_a.path, source_b=candidate.source_b.path)
        result.add_change(ChangeRecord(key="line_1", change_type=ChangeType.MODIFIED))
        return result

    monkeypatch.setattr(
        "src.services.comparison.drawing_batch.compare_candidate",
        fake_compare,
    )

    state_dir = tmp_path / "state"
    summary = BatchCompareJob(
        [candidate],
        options=BatchCompareOptions(
            compare_state_dir=state_dir,
            write_compare_state_json=False,
            max_workers=1,
        ),
    ).run()

    assert summary.completed_pairs == 1
    assert not (state_dir / "compare_state.json").exists()


def test_batch_job_forwards_inner_cad_progress(monkeypatch) -> None:
    events: list[tuple[int, int, str]] = []

    class FakeDwgDiffer:
        def __init__(self, **_kwargs) -> None:
            pass

        def compare(
            self,
            source_a,
            source_b,
            *,
            progress_callback=None,
            is_cancelled=None,
        ) -> ComparisonResult:
            if progress_callback:
                progress_callback(80, 100, "inner CAD compare")
            return ComparisonResult(source_a=str(source_a), source_b=str(source_b))

    monkeypatch.setattr("src.services.comparison.dwg_differ.DwgDiffer", FakeDwgDiffer)

    summary = BatchCompareJob([_confirmed_candidate("S-101")]).run(
        progress_callback=lambda current, total, message: events.append(
            (current, total, message)
        )
    )

    assert summary.completed_pairs == 1
    assert (80, 100, "inner CAD compare") in events


def test_cad_batch_compare_uses_canonical_engine_by_default(monkeypatch) -> None:
    captured: dict[str, object] = {}

    class FakeDwgDiffer:
        def __init__(self, **kwargs) -> None:
            captured.update(kwargs)

        def compare(
            self,
            source_a,
            source_b,
            *,
            progress_callback=None,
            is_cancelled=None,
        ) -> ComparisonResult:
            return ComparisonResult(source_a=str(source_a), source_b=str(source_b))

    monkeypatch.setattr("src.services.comparison.dwg_differ.DwgDiffer", FakeDwgDiffer)

    result = compare_candidate(_confirmed_candidate("S-101"), BatchCompareOptions())

    assert result.source_a.endswith("S-101.dwg")
    assert captured["config"] == {
        "use_canonical_pipeline": True,
        "use_legacy_ezdxf_pipeline": False,
    }


def test_cad_batch_compare_can_opt_into_legacy_ezdxf_engine(monkeypatch) -> None:
    captured: dict[str, object] = {}

    class FakeDwgDiffer:
        def __init__(self, **kwargs) -> None:
            captured.update(kwargs)

        def compare(
            self,
            source_a,
            source_b,
            *,
            progress_callback=None,
            is_cancelled=None,
        ) -> ComparisonResult:
            return ComparisonResult(source_a=str(source_a), source_b=str(source_b))

    monkeypatch.setattr("src.services.comparison.dwg_differ.DwgDiffer", FakeDwgDiffer)

    compare_candidate(
        _confirmed_candidate("S-101"),
        BatchCompareOptions(cad_compare_engine="legacy_ezdxf"),
    )

    assert captured["config"] == {
        "use_canonical_pipeline": False,
        "use_legacy_ezdxf_pipeline": True,
    }


def test_cad_batch_compare_falls_back_to_legacy_ezdxf_for_failed_cached_dxf(monkeypatch) -> None:
    calls: list[dict[str, object]] = []
    candidate = MatchCandidate(
        source_a=_descriptor("S-101.dxf"),
        source_b=_descriptor("S-101_REV1.dxf"),
        score=0.9,
        status=MatchStatus.AUTO_CONFIRMED,
    )

    class FakeDwgDiffer:
        def __init__(self, **kwargs) -> None:
            self.kwargs = kwargs
            calls.append(kwargs)

        def compare(
            self,
            source_a,
            source_b,
            *,
            progress_callback=None,
            is_cancelled=None,
        ) -> ComparisonResult:
            result = ComparisonResult(source_a=str(source_a), source_b=str(source_b))
            if self.kwargs["config"]["use_canonical_pipeline"]:
                result.metadata.update(
                    {
                        "pipeline_status": "failed",
                        "error_code": "COMPARE_IMPORT_FAILED",
                        "message": "CAD compare import failed",
                    }
                )
                return result
            result.add_change(ChangeRecord(key="line_1", change_type=ChangeType.ADDED))
            result.metadata["comparison_type"] = "DWG/DXF"
            return result

    monkeypatch.setattr("src.services.comparison.dwg_differ.DwgDiffer", FakeDwgDiffer)

    result = compare_candidate(candidate, BatchCompareOptions())

    assert result.total_changes == 1
    assert result.metadata["canonical_fallback_used"] is True
    assert result.metadata["canonical_error_code"] == "COMPARE_IMPORT_FAILED"
    assert [call["config"]["use_canonical_pipeline"] for call in calls] == [True, False]


def test_large_dxf_batch_compare_preselects_legacy_ezdxf(
    tmp_path: Path,
    monkeypatch,
) -> None:
    source_a = tmp_path / "S-101.dxf"
    source_b = tmp_path / "S-101_REV1.dxf"
    source_a.write_bytes(b"0\nSECTION\n")
    source_b.write_bytes(b"0\nSECTION\n")
    calls: list[dict[str, object]] = []

    def descriptor_for(path: Path) -> DrawingFileDescriptor:
        return DrawingFileDescriptor(
            path=str(path),
            kind=DrawingKind.CAD,
            extension=".dxf",
            relative_path=path.name,
            identity=parse_filename_identity(path.name),
        )

    candidate = MatchCandidate(
        source_a=descriptor_for(source_a),
        source_b=descriptor_for(source_b),
        score=0.9,
        status=MatchStatus.AUTO_CONFIRMED,
    )

    class FakeDwgDiffer:
        def __init__(self, **kwargs) -> None:
            self.kwargs = kwargs
            calls.append(kwargs)

        def compare(
            self,
            source_a,
            source_b,
            *,
            progress_callback=None,
            is_cancelled=None,
        ) -> ComparisonResult:
            result = ComparisonResult(source_a=str(source_a), source_b=str(source_b))
            result.add_change(ChangeRecord(key="line_1", change_type=ChangeType.ADDED))
            return result

    monkeypatch.setenv("DRAWING_COMPARE_LEGACY_DXF_DIRECT_MB", "0.000001")
    monkeypatch.setattr("src.services.comparison.dwg_differ.DwgDiffer", FakeDwgDiffer)

    result = compare_candidate(candidate, BatchCompareOptions())

    assert result.total_changes == 1
    assert result.metadata["legacy_ezdxf_preselected"] is True
    assert result.warnings
    assert len(calls) == 1
    assert calls[0]["config"] == {
        "use_canonical_pipeline": False,
        "use_legacy_ezdxf_pipeline": True,
    }


def test_legacy_ezdxf_fallback_available_accepts_same_stem_cache(tmp_path: Path) -> None:
    cache_dir = tmp_path / "dxf_cache"
    cache_dir.mkdir()
    source_a = tmp_path / "source_a" / "large_detail.dwg"
    source_b = tmp_path / "source_b" / "large_detail_R1.dwg"
    source_a.parent.mkdir()
    source_b.parent.mkdir()
    source_a.write_bytes(b"dwg-a")
    source_b.write_bytes(b"dwg-b")
    (cache_dir / "large_detail.oldhash.dxf").write_text("0\nEOF\n", encoding="utf-8")
    (cache_dir / "large_detail_R1.oldhash.dxf").write_text("0\nEOF\n", encoding="utf-8")

    assert _legacy_ezdxf_fallback_available(source_a, source_b, cache_dir) is True


def test_pair_uuid_distinguishes_same_label_in_different_folders() -> None:
    first_a = _descriptor("S-101.dwg", drawing_number="S101")
    first_b = _descriptor("S-101_REV1.dwg", drawing_number="S101")
    second_a = _descriptor("S-101.dwg", drawing_number="S101")
    second_b = _descriptor("S-101_REV1.dwg", drawing_number="S101")
    first_a.path = "C:/project/structural/A/S-101.dwg"
    first_b.path = "C:/project/structural/B/S-101_REV1.dwg"
    second_a.path = "C:/project/architectural/A/S-101.dwg"
    second_b.path = "C:/project/architectural/B/S-101_REV1.dwg"

    first = MatchCandidate(first_a, first_b, score=0.95, status=MatchStatus.AUTO_CONFIRMED)
    second = MatchCandidate(second_a, second_b, score=0.95, status=MatchStatus.AUTO_CONFIRMED)

    assert candidate_pair_uuid(first) != candidate_pair_uuid(second)
    assert first.to_dict()["pair_uuid"] == candidate_pair_uuid(first)
    assert first.to_dict()["display_label"] == "S101"


def test_confirmed_pairs_must_be_one_to_one_for_a_and_b() -> None:
    source_a = _descriptor("S-101.dwg")
    first_b = _descriptor("S-101_REV1.dwg")
    second_b = _descriptor("S-101_COPY_REV2.dwg")
    first = MatchCandidate(source_a, first_b, score=0.95, status=MatchStatus.MANUAL_CONFIRMED)
    second = MatchCandidate(source_a, second_b, score=0.95, status=MatchStatus.MANUAL_CONFIRMED)

    violations = confirmed_pair_uniqueness_violations([first, second])

    assert violations["duplicate_a"] == [source_a.path]
    with pytest.raises(ValueError, match="one-to-one"):
        BatchCompareJob([first, second]).run()


def test_batch_summary_and_exports_use_full_metadata_change_counts(tmp_path) -> None:
    candidate = _confirmed_candidate("S-201")
    result = ComparisonResult(source_a=candidate.source_a.path, source_b=candidate.source_b.path)
    result.add_change(ChangeRecord(key="line_1", change_type=ChangeType.DELETED))
    result.add_change(ChangeRecord(key="line_2", change_type=ChangeType.DELETED))
    result.metadata["change_counts"] = {"added": 3, "deleted": 2, "modified": 1}
    result.metadata["truncated_changes"] = True

    summary = BatchCompareSummary(
        started_at=datetime.now(),
        requested_pairs=1,
        items=[
            BatchCompareItemResult(
                candidate=candidate,
                result=result,
                status="completed",
            )
        ],
    )

    assert result.total_changes == 2
    assert summary.total_changes == 6

    html_path = BatchCompareJob.export_html(summary, tmp_path / "report.html")
    html = html_path.read_text(encoding="utf-8")
    assert "total changes: 6" in html
    assert "<td>6</td>" in html

    xlsx_path = BatchCompareJob.export_excel(summary, tmp_path / "report.xlsx")
    import openpyxl

    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb["Summary"]
    assert ws.cell(row=2, column=6).value == 6
    run_ws = wb["Run Summary"]
    assert run_ws.cell(row=7, column=2).value == 6


def test_batch_job_records_cancelled_pairs_in_sequential_mode(monkeypatch) -> None:
    candidates = [_confirmed_candidate(f"S-{200 + index}") for index in range(3)]
    cancelled = {"value": False}

    def fake_compare(candidate, options, is_cancelled=None):
        result = ComparisonResult(source_a=candidate.source_a.path, source_b=candidate.source_b.path)
        result.add_change(ChangeRecord(key="line_1", change_type=ChangeType.MODIFIED))
        cancelled["value"] = True
        return result

    monkeypatch.setattr(
        "src.services.comparison.drawing_batch.compare_candidate",
        fake_compare,
    )

    summary = BatchCompareJob(candidates, options=BatchCompareOptions(max_workers=1)).run(
        is_cancelled=lambda: cancelled["value"],
    )

    assert summary.requested_pairs == 3
    assert summary.total_pairs == 3
    assert summary.completed_pairs == 1
    assert summary.cancelled_pairs == 2
    assert summary.cancelled is True
    assert [item.status for item in summary.items] == ["completed", "cancelled", "cancelled"]


def test_batch_job_records_cancelled_pairs_in_parallel_mode() -> None:
    candidates = [_confirmed_candidate(f"S-{300 + index}") for index in range(3)]

    summary = BatchCompareJob(candidates, options=BatchCompareOptions(max_workers=2)).run(
        is_cancelled=lambda: True,
    )

    assert summary.requested_pairs == 3
    assert summary.total_pairs == 3
    assert summary.completed_pairs == 0
    assert summary.cancelled_pairs == 3
    assert summary.cancelled is True


def test_descriptor_cache_can_be_disabled(tmp_path) -> None:
    root = tmp_path / "drawings"
    root.mkdir()
    (root / "A-100.dxf").write_text("invalid dxf placeholder", encoding="utf-8")

    descriptors = scan_drawing_inputs(root, DescriptorBuildOptions(enable_cache=False))

    assert [descriptor.name for descriptor in descriptors] == ["A-100.dxf"]
    assert not (root / ".drawing_compare_cache").exists()


def test_descriptor_cache_key_includes_extraction_options(tmp_path) -> None:
    root = tmp_path / "drawings"
    root.mkdir()
    (root / "A-100.dxf").write_text("invalid dxf placeholder", encoding="utf-8")

    scan_drawing_inputs(root, DescriptorBuildOptions(enable_cache=True, max_text_chars=4000))
    cache_files = sorted((root / ".drawing_compare_cache").glob("*.json"))
    assert len(cache_files) == 1

    scan_drawing_inputs(root, DescriptorBuildOptions(enable_cache=True, max_text_chars=120))
    cache_files = sorted((root / ".drawing_compare_cache").glob("*.json"))
    assert len(cache_files) == 2


def test_corrupt_descriptor_cache_is_ignored_and_rewritten(tmp_path) -> None:
    root = tmp_path / "drawings"
    root.mkdir()
    (root / "A-100.dxf").write_text("invalid dxf placeholder", encoding="utf-8")

    scan_drawing_inputs(root, DescriptorBuildOptions(enable_cache=True))
    cache_file = next((root / ".drawing_compare_cache").glob("*.json"))
    cache_file.write_text("{not valid json", encoding="utf-8")

    descriptors = scan_drawing_inputs(root, DescriptorBuildOptions(enable_cache=True))

    assert [descriptor.name for descriptor in descriptors] == ["A-100.dxf"]
    assert "cache_version" in cache_file.read_text(encoding="utf-8")


def test_manual_review_csv_round_trips_and_applies_statuses(tmp_path) -> None:
    old = _descriptor("S-501_REV0.dwg")
    first_b = _descriptor("S-501_REV1.dwg")
    second_b = _descriptor("S-501_COPY_REV2.dwg")
    candidates = match_drawing_sets([old], [first_b, second_b])
    review_path = tmp_path / "review.csv"

    write_manual_match_csv(candidates, review_path)
    rows = review_path.read_text(encoding="utf-8-sig")
    assert "a_path,b_path,status" in rows
    assert "review_required" in rows

    result = apply_manual_matches(
        [{"a_path": old.path, "b_path": first_b.path, "status": "manual_confirmed"}],
        candidates,
        [old],
        [first_b, second_b],
    )

    assert result["applied"] == 1
    assert any(candidate.status == MatchStatus.MANUAL_CONFIRMED for candidate in candidates)
