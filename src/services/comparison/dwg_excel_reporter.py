"""DWG/DXF 비교 결과 Excel 리포트 생성

Sprint 10 Feature 4: DwgExcelReporter
비교 결과를 Excel 파일로 내보냅니다.

시트 구성:
    - 요약: 파일 정보 및 통계
    - 변경 목록: 상세 변경 사항
    - 레이어별 통계: 레이어별 추가/삭제 수
    - 엔티티별 통계: 엔티티 타입별 추가/삭제 수
"""

import logging
from datetime import datetime
from pathlib import Path
from typing import List, Optional

logger = logging.getLogger(__name__)

# openpyxl 임포트 (선택적)
try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter

    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    logger.warning("openpyxl이 설치되지 않았습니다: pip install openpyxl")

from .dxf_comparator import DxfChange, DxfChangeType, DxfComparisonResult


class DwgExcelReporter:
    """DWG 비교 결과 Excel 리포트 생성기

    사용 예시:
        reporter = DwgExcelReporter()
        reporter.generate(result, "old.dwg", "new.dwg", Path("report.xlsx"))
    """

    # 스타일 정의
    HEADER_FILL = None
    HEADER_FONT = None
    ADDED_FILL = None
    DELETED_FILL = None
    MODIFIED_FILL = None
    THIN_BORDER = None

    def __init__(self):
        if not OPENPYXL_AVAILABLE:
            raise ImportError("openpyxl이 필요합니다: pip install openpyxl")

        # 스타일 초기화
        self.HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
        self.HEADER_FONT = Font(color="FFFFFF", bold=True)
        self.ADDED_FILL = PatternFill("solid", fgColor="C6EFCE")
        self.DELETED_FILL = PatternFill("solid", fgColor="FFC7CE")
        self.MODIFIED_FILL = PatternFill("solid", fgColor="FFEB9C")
        self.THIN_BORDER = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

    def generate(
        self,
        result: DxfComparisonResult,
        source_a: str,
        source_b: str,
        output_path: Path,
    ) -> Path:
        """Excel 리포트 생성

        Args:
            result: DXF 비교 결과
            source_a: 기준 파일 경로
            source_b: 대상 파일 경로
            output_path: 출력 파일 경로

        Returns:
            생성된 Excel 파일 경로
        """
        output_path = Path(output_path)

        wb = openpyxl.Workbook()

        # 1. 요약 시트
        self._create_summary_sheet(wb.active, result, source_a, source_b)

        # 2. 변경 목록 시트
        ws_changes = wb.create_sheet("변경 목록")
        self._create_changes_sheet(ws_changes, result.changes)

        # 3. 레이어별 통계 시트
        ws_layers = wb.create_sheet("레이어별 통계")
        self._create_layer_stats_sheet(ws_layers, result)

        # 4. 엔티티별 통계 시트
        ws_types = wb.create_sheet("엔티티별 통계")
        self._create_type_stats_sheet(ws_types, result)

        # 5. 치수 변경 시트 (치수 변경이 있는 경우)
        dim_changes = [c for c in result.changes if c.is_dimension_change()]
        if dim_changes:
            ws_dims = wb.create_sheet("치수 변경")
            self._create_dimension_sheet(ws_dims, dim_changes)

        # 저장
        wb.save(str(output_path))

        logger.info(f"Excel 리포트 생성: {output_path}")
        return output_path

    def _create_summary_sheet(self, ws, result: DxfComparisonResult, source_a: str, source_b: str):
        """요약 시트 생성"""
        ws.title = "요약"

        data = [
            ["DWG/DXF 비교 리포트"],
            [],
            ["생성 일시", datetime.now().strftime("%Y-%m-%d %H:%M:%S")],
            ["기준 파일 (Old)", source_a],
            ["대상 파일 (New)", source_b],
            [],
            ["📊 비교 결과"],
            ["총 변경점", result.total_changes],
            ["추가", result.added_count],
            ["삭제", result.deleted_count],
            ["수정", result.modified_count],
            [],
            ["📁 엔티티 수"],
            ["기준 파일 엔티티", result.stats.get("entities_a", 0)],
            ["대상 파일 엔티티", result.stats.get("entities_b", 0)],
        ]

        for row_data in data:
            ws.append(row_data)

        # 스타일 적용
        ws["A1"].font = Font(size=16, bold=True)
        ws["A7"].font = Font(size=12, bold=True)
        ws["A13"].font = Font(size=12, bold=True)

        ws.column_dimensions["A"].width = 25
        ws.column_dimensions["B"].width = 60

    def _create_changes_sheet(self, ws, changes: List[DxfChange]):
        """변경 목록 시트 생성"""
        headers = ["#", "변경유형", "엔티티", "레이어", "위치 X", "위치 Y", "상세정보"]
        ws.append(headers)

        # 헤더 스타일
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col)
            cell.fill = self.HEADER_FILL
            cell.font = self.HEADER_FONT
            cell.border = self.THIN_BORDER

        # 데이터 추가
        for i, change in enumerate(changes, 1):
            loc_x = f"{change.location[0]:.1f}" if change.location else ""
            loc_y = f"{change.location[1]:.1f}" if change.location else ""

            # 상세 정보
            detail = ""
            if change.is_dimension_change():
                meas_change = change.get_measurement_change()
                if meas_change:
                    detail = f"치수: {meas_change}"
            else:
                data = change.new_data or change.old_data or {}
                detail = str(data)[:50] + "..." if len(str(data)) > 50 else str(data)

            row = [
                i,
                change.change_type.value,
                change.entity_type,
                change.layer,
                loc_x,
                loc_y,
                detail,
            ]
            ws.append(row)

            # 행 색상
            row_num = i + 1
            fill = None
            if change.change_type == DxfChangeType.ADDED:
                fill = self.ADDED_FILL
            elif change.change_type == DxfChangeType.DELETED:
                fill = self.DELETED_FILL
            elif change.change_type == DxfChangeType.MODIFIED:
                fill = self.MODIFIED_FILL

            if fill:
                for col in range(1, len(headers) + 1):
                    ws.cell(row=row_num, column=col).fill = fill

        # 열 너비 조정
        widths = [5, 10, 12, 20, 12, 12, 40]
        for col, width in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(col)].width = width

    def _create_layer_stats_sheet(self, ws, result: DxfComparisonResult):
        """레이어별 통계 시트"""
        headers = ["레이어", "추가", "삭제", "합계"]
        ws.append(headers)

        # 헤더 스타일
        for col in range(1, 5):
            cell = ws.cell(row=1, column=col)
            cell.fill = self.HEADER_FILL
            cell.font = self.HEADER_FONT

        # 레이어별 집계
        by_layer = result.stats.get("by_layer", {})
        for layer, stats in sorted(by_layer.items()):
            added = stats.get("added", 0)
            deleted = stats.get("deleted", 0)
            ws.append([layer, added, deleted, added + deleted])

        # 열 너비
        ws.column_dimensions["A"].width = 30
        for col in ["B", "C", "D"]:
            ws.column_dimensions[col].width = 10

    def _create_type_stats_sheet(self, ws, result: DxfComparisonResult):
        """엔티티별 통계 시트"""
        headers = ["엔티티 타입", "추가", "삭제", "합계"]
        ws.append(headers)

        for col in range(1, 5):
            cell = ws.cell(row=1, column=col)
            cell.fill = self.HEADER_FILL
            cell.font = self.HEADER_FONT

        by_type = result.stats.get("by_type", {})
        for etype, stats in sorted(by_type.items()):
            added = stats.get("added", 0)
            deleted = stats.get("deleted", 0)
            ws.append([etype, added, deleted, added + deleted])

        ws.column_dimensions["A"].width = 20
        for col in ["B", "C", "D"]:
            ws.column_dimensions[col].width = 10

    def _create_dimension_sheet(self, ws, dim_changes: List[DxfChange]):
        """치수 변경 시트"""
        headers = ["#", "레이어", "위치 X", "위치 Y", "이전값", "신규값", "변화량"]
        ws.append(headers)

        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col)
            cell.fill = self.HEADER_FILL
            cell.font = self.HEADER_FONT

        for i, change in enumerate(dim_changes, 1):
            loc_x = f"{change.location[0]:.1f}" if change.location else ""
            loc_y = f"{change.location[1]:.1f}" if change.location else ""

            old_val = (change.old_data or {}).get("measurement", "-")
            new_val = (change.new_data or {}).get("measurement", "-")

            diff = ""
            if isinstance(old_val, (int, float)) and isinstance(new_val, (int, float)):
                d = new_val - old_val
                diff = f"{'+' if d > 0 else ''}{d:.1f}"

            ws.append([i, change.layer, loc_x, loc_y, old_val, new_val, diff])

        # 열 너비
        widths = [5, 20, 12, 12, 12, 12, 12]
        for col, width in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(col)].width = width
