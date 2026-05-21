# -*- coding: utf-8 -*-
"""
엑셀 비교 엔진 (Excel Differ)
=============================

Key Column 기반으로 두 엑셀 파일의 변경 내역을 추적합니다.

주요 기능:
- 사용자 지정 Key Column 기반 Row 매칭
- Numeric Tolerance 지원 (부동소수점 오차 무시)
- Format Ignore (1,000 vs 1000 동일 처리)
- Added/Deleted/Modified Row 분류

Author: TEKLA_MCP Team
Date: 2025-12-14
"""

import logging
from contextlib import contextmanager
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Set, Tuple, Union
import pandas as pd

from .base import (
    BaseComparator,
    ChangeRecord,
    ChangeType,
    ComparisonResult,
)

logger = logging.getLogger(__name__)


class ExcelDiffer(BaseComparator):
    """Key 기반 엑셀 비교 엔진

    Examples:
        >>> differ = ExcelDiffer(config={
        ...     "key_columns": ["부재번호", "층"],
        ...     "numeric_tolerance": 0.001,
        ...     "ignore_columns": ["비고"],
        ... })
        >>> result = differ.compare("old.xlsx", "new.xlsx")
        >>> print(f"변경: {result.total_changes}건")
    """

    def __init__(self, config: Optional[Dict[str, Any]] = None):
        super().__init__(config)

        # 기본 설정
        self._key_columns: List[str] = self.get_config("key_columns", [])
        self._numeric_tolerance: float = self.get_config("numeric_tolerance", 0.001)
        self._ignore_columns: Set[str] = set(self.get_config("ignore_columns", []))
        self._sheet_name: Optional[str] = self.get_config("sheet_name", None)

    def compare(
        self,
        source_a: Union[str, Path],
        source_b: Union[str, Path],
    ) -> ComparisonResult:
        """두 엑셀 파일을 비교합니다."""
        source_a = Path(source_a)
        source_b = Path(source_b)

        logger.info(f"Excel 비교 시작: {source_a.name} vs {source_b.name}")

        # 결과 객체 생성
        self._result = ComparisonResult(
            source_a=str(source_a),
            source_b=str(source_b),
        )

        try:
            # 엑셀 파일 로드
            df_a = self._load_excel(source_a)
            df_b = self._load_excel(source_b)

            # Key Column 자동 감지 (미지정 시)
            if not self._key_columns:
                self._key_columns = self._auto_detect_key_columns(df_a, df_b)
                self._result.warnings.append(f"Key Column 자동 감지: {self._key_columns}")

            # Key Column 유효성 검사
            self._validate_key_columns(df_a, df_b)

            # 비교 수행
            self._compare_dataframes(df_a, df_b)

            # 메타데이터 추가
            self._result.metadata = {
                "key_columns": self._key_columns,
                "numeric_tolerance": self._numeric_tolerance,
                "ignore_columns": list(self._ignore_columns),
                "rows_in_a": len(df_a),
                "rows_in_b": len(df_b),
            }

            logger.info(
                f"Excel 비교 완료: "
                f"추가 {self._result.added_count}, "
                f"삭제 {self._result.deleted_count}, "
                f"수정 {self._result.modified_count}"
            )

        except Exception as e:
            logger.error(f"Excel 비교 실패: {e}")
            self._result.warnings.append(f"비교 중 오류 발생: {e}")
            raise

        return self._result

    def compare_large(
        self,
        source_a: Union[str, Path],
        source_b: Union[str, Path],
        chunksize: int = 10000,
    ) -> ComparisonResult:
        """대용량 엑셀 파일을 Chunked Processing으로 비교합니다.

        메모리 효율적인 비교를 위해 파일을 청크 단위로 로드하여 처리합니다.

        Args:
            source_a: 비교 기준 파일 (Old)
            source_b: 비교 대상 파일 (New)
            chunksize: 한 번에 로드할 Row 수 (기본값: 10,000)

        Returns:
            ComparisonResult: 비교 결과

        Note:
            - 대용량 파일(10만 행 이상)에 권장
            - 메모리 사용량: 약 chunksize * 2 * row_size
        """
        source_a = Path(source_a)
        source_b = Path(source_b)

        logger.info(f"[Large File Mode] Excel 비교 시작: {source_a.name} vs {source_b.name}")
        logger.info(f"  Chunk size: {chunksize:,} rows")

        # 결과 객체 생성
        self._result = ComparisonResult(
            source_a=str(source_a),
            source_b=str(source_b),
        )

        try:
            # Key Column이 지정되지 않은 경우, 첫 청크에서 감지
            if not self._key_columns:
                sample_a = pd.read_excel(source_a, nrows=100)
                sample_b = pd.read_excel(source_b, nrows=100)
                self._key_columns = self._auto_detect_key_columns(sample_a, sample_b)
                self._result.warnings.append(f"Key Column 자동 감지: {self._key_columns}")

            # 전체 Key 수집 (첫 번째 패스)
            logger.info("  [Pass 1] 전체 Key 수집 중...")
            keys_a, dup_a = self._collect_keys_chunked(source_a, chunksize)
            keys_b, dup_b = self._collect_keys_chunked(source_b, chunksize)

            if dup_a:
                self._result.warnings.append(
                    f"파일 A에 중복된 Key가 {dup_a}건 존재합니다. 첫 번째 항목만 비교됩니다."
                )
            if dup_b:
                self._result.warnings.append(
                    f"파일 B에 중복된 Key가 {dup_b}건 존재합니다. 첫 번째 항목만 비교됩니다."
                )

            # 삭제/추가된 키 식별
            deleted_keys = keys_a - keys_b
            added_keys = keys_b - keys_a
            common_keys = keys_a & keys_b

            logger.info(f"    Keys in A: {len(keys_a):,}, Keys in B: {len(keys_b):,}")
            logger.info(
                f"    Added: {len(added_keys):,}, Deleted: {len(deleted_keys):,}, Common: {len(common_keys):,}"
            )

            # 삭제된 항목 기록
            for key in deleted_keys:
                self._result.add_change(
                    ChangeRecord(
                        key=str(key),
                        change_type=ChangeType.DELETED,
                        location=f"Row (Key: {key})",
                    )
                )

            # 추가된 항목 기록
            for key in added_keys:
                self._result.add_change(
                    ChangeRecord(
                        key=str(key),
                        change_type=ChangeType.ADDED,
                        location=f"Row (Key: {key})",
                    )
                )

            # 공통 키에 대해 청크 단위로 비교 (두 번째 패스)
            if common_keys:
                logger.info("  [Pass 2] 공통 Key 비교 중...")
                self._compare_common_keys_chunked(source_a, source_b, common_keys, chunksize)

            # 메타데이터 추가
            self._result.metadata = {
                "key_columns": self._key_columns,
                "numeric_tolerance": self._numeric_tolerance,
                "ignore_columns": list(self._ignore_columns),
                "mode": "chunked",
                "chunksize": chunksize,
                "total_keys_a": len(keys_a),
                "total_keys_b": len(keys_b),
            }

            logger.info(
                f"[Large File Mode] 비교 완료: "
                f"추가 {self._result.added_count}, "
                f"삭제 {self._result.deleted_count}, "
                f"수정 {self._result.modified_count}"
            )

        except Exception as e:
            logger.error(f"Excel 비교 실패: {e}")
            self._result.warnings.append(f"비교 중 오류 발생: {e}")
            raise

        return self._result

    def _collect_keys_chunked(self, path: Path, chunksize: int) -> Tuple[Set, int]:
        """Key 수집 (메모리 효율적)

        Note:
            .xlsx 계열은 openpyxl streaming으로 Key 컬럼만 순회합니다.
            그 외 확장자는 pandas로 fallback 합니다.
        """
        if self._can_stream_excel(path):
            with self._openpyxl_reader(path) as (header, rows):
                if not header:
                    return set(), 0
                key_indices = self._get_key_indices(header)
                keys: Set = set()
                dup_count = 0
                for row in rows:
                    if not row or all(value is None for value in row):
                        continue
                    key = self._make_key_from_row(row, key_indices)
                    if key in keys:
                        dup_count += 1
                        continue
                    keys.add(key)
                return keys, dup_count

        # Fallback: pandas
        try:
            df = pd.read_excel(path, usecols=self._key_columns, sheet_name=self._sheet_name or 0)
        except Exception:
            df = pd.read_excel(path, sheet_name=self._sheet_name or 0)

        df.columns = [str(c).strip() for c in df.columns]

        if len(self._key_columns) == 1:
            series = df[self._key_columns[0]].astype(str)
            dup_count = int(series.duplicated().sum())
            return set(series), dup_count

        key_series = df[self._key_columns].apply(tuple, axis=1)
        dup_count = int(key_series.duplicated().sum())
        return set(key_series), dup_count

    def _compare_common_keys_chunked(
        self,
        path_a: Path,
        path_b: Path,
        common_keys: Set,
        chunksize: int,
    ) -> None:
        """공통 Key에 대해 비교 (메모리 효율적)

        공통 Key만 필터링하여 처리합니다.
        """
        if self._can_stream_excel(path_a) and self._can_stream_excel(path_b):
            header_a = self._read_header(path_a)
            header_b = self._read_header(path_b)
            if not header_a or not header_b:
                return

            compare_cols = self._build_compare_columns(header_a, header_b)
            if not compare_cols:
                return

            key_indices_a = self._get_key_indices(header_a)
            key_indices_b = self._get_key_indices(header_b)

            col_indices_a = {col: header_a.index(col) for col in compare_cols}
            col_indices_b = {col: header_b.index(col) for col in compare_cols}

            data_b: Dict[Any, Dict[str, Any]] = {}
            with self._openpyxl_reader(path_b) as (_, rows_b):
                for row in rows_b:
                    if not row or all(value is None for value in row):
                        continue
                    key = self._make_key_from_row(row, key_indices_b)
                    if key not in common_keys:
                        continue
                    if key in data_b:
                        continue
                    data_b[key] = {
                        col: self._get_row_value(row, col_indices_b[col]) for col in compare_cols
                    }

            seen_a: Set[Any] = set()
            with self._openpyxl_reader(path_a) as (_, rows_a):
                for row in rows_a:
                    if not row or all(value is None for value in row):
                        continue
                    key = self._make_key_from_row(row, key_indices_a)
                    if key not in common_keys or key in seen_a:
                        continue
                    seen_a.add(key)

                    row_b = data_b.get(key)
                    if row_b is None:
                        continue

                    for col in compare_cols:
                        val_a = self._get_row_value(row, col_indices_a[col])
                        val_b = row_b.get(col)
                        if not self._values_equal(val_a, val_b):
                            self._result.add_change(
                                ChangeRecord(
                                    key=str(key),
                                    change_type=ChangeType.MODIFIED,
                                    field_name=col,
                                    old_value=val_a,
                                    new_value=val_b,
                                    location=f"Row (Key: {key}), Column: {col}",
                                )
                            )
            return

        # Fallback: pandas
        df_b = pd.read_excel(path_b, sheet_name=self._sheet_name or 0)
        df_b.columns = [str(c).strip() for c in df_b.columns]

        if len(self._key_columns) == 1:
            df_b["_key"] = df_b[self._key_columns[0]].astype(str)
        else:
            df_b["_key"] = df_b[self._key_columns].apply(tuple, axis=1)

        if df_b["_key"].duplicated().any():
            df_b = df_b.loc[~df_b["_key"].duplicated(keep="first")]

        df_b_filtered = df_b[df_b["_key"].isin(common_keys)].set_index("_key")
        data_b = df_b_filtered.to_dict("index")
        del df_b, df_b_filtered

        df_a = pd.read_excel(path_a, sheet_name=self._sheet_name or 0)
        df_a.columns = [str(c).strip() for c in df_a.columns]

        if len(self._key_columns) == 1:
            df_a["_key"] = df_a[self._key_columns[0]].astype(str)
        else:
            df_a["_key"] = df_a[self._key_columns].apply(tuple, axis=1)

        all_cols = set(df_a.columns)
        compare_cols = list(all_cols - set(self._key_columns) - self._ignore_columns - {"_key"})

        df_a_filtered = df_a[df_a["_key"].isin(common_keys)]
        if df_a_filtered["_key"].duplicated().any():
            df_a_filtered = df_a_filtered.loc[~df_a_filtered["_key"].duplicated(keep="first")]

        for _, row in df_a_filtered.iterrows():
            key = row["_key"]
            row_b = data_b.get(key)

            if not row_b:
                continue

            for col in compare_cols:
                val_a = row.get(col)
                val_b = row_b.get(col)

                if not self._values_equal(val_a, val_b):
                    self._result.add_change(
                        ChangeRecord(
                            key=str(key),
                            change_type=ChangeType.MODIFIED,
                            field_name=col,
                            old_value=val_a,
                            new_value=val_b,
                            location=f"Row (Key: {key}), Column: {col}",
                        )
                    )

    def _can_stream_excel(self, path: Path) -> bool:
        return path.suffix.lower() in {".xlsx", ".xlsm", ".xltx", ".xltm"}

    @contextmanager
    def _openpyxl_reader(self, path: Path) -> Iterable[Tuple[List[str], Iterable[Tuple[Any, ...]]]]:
        import openpyxl

        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        try:
            if self._sheet_name is None:
                ws = wb.active
            elif isinstance(self._sheet_name, int):
                ws = wb.worksheets[self._sheet_name]
            else:
                ws = wb[self._sheet_name]

            rows = ws.iter_rows(values_only=True)
            header = next(rows, None)
            if header is None:
                yield [], iter(())
                return
            header = [str(c).strip() if c is not None else "" for c in header]
            yield header, rows
        finally:
            wb.close()

    def _read_header(self, path: Path) -> List[str]:
        with self._openpyxl_reader(path) as (header, _):
            return header

    def _get_key_indices(self, header: List[str]) -> List[int]:
        indices = []
        for key in self._key_columns:
            if key not in header:
                raise ValueError(f"Key Column '{key}'이 파일에 없습니다.")
            indices.append(header.index(key))
        return indices

    def _build_compare_columns(self, header_a: List[str], header_b: List[str]) -> List[str]:
        shared = set(header_a) & set(header_b)
        shared.discard("")
        compare_cols = list(shared - set(self._key_columns) - self._ignore_columns)
        return compare_cols

    def _make_key_from_row(self, row: Tuple[Any, ...], key_indices: List[int]) -> Any:
        if len(key_indices) == 1:
            value = self._get_row_value(row, key_indices[0])
            return "nan" if value is None else str(value)
        return tuple(self._get_row_value(row, idx) for idx in key_indices)

    def _get_row_value(self, row: Tuple[Any, ...], idx: int) -> Any:
        if idx < len(row):
            return row[idx]
        return None

    def _load_excel(self, path: Path) -> pd.DataFrame:
        """엑셀 파일 로드"""
        try:
            if self._sheet_name:
                df = pd.read_excel(path, sheet_name=self._sheet_name)
            else:
                df = pd.read_excel(path)

            # 컬럼명 정규화 (공백 제거)
            df.columns = [str(c).strip() for c in df.columns]

            return df
        except Exception as e:
            raise ValueError(f"엑셀 파일 로드 실패 ({path}): {e}")

    def _auto_detect_key_columns(
        self,
        df_a: pd.DataFrame,
        df_b: pd.DataFrame,
    ) -> List[str]:
        """Key Column 자동 감지"""
        common_cols = set(df_a.columns) & set(df_b.columns)

        # ID/번호 관련 컬럼 우선 탐색
        key_patterns = ["id", "번호", "no", "key", "code", "부재", "요소"]

        for col in common_cols:
            col_lower = col.lower()
            if any(p in col_lower for p in key_patterns):
                return [col]

        # 첫 번째 컬럼을 Key로 사용
        if common_cols:
            first_col = list(df_a.columns)[0]
            if first_col in common_cols:
                return [first_col]

        raise ValueError("Key Column을 자동으로 감지할 수 없습니다. 명시적으로 지정해주세요.")

    def _validate_key_columns(
        self,
        df_a: pd.DataFrame,
        df_b: pd.DataFrame,
    ) -> None:
        """Key Column 유효성 검사"""
        for key_col in self._key_columns:
            if key_col not in df_a.columns:
                raise ValueError(f"Key Column '{key_col}'이 파일 A에 없습니다.")
            if key_col not in df_b.columns:
                raise ValueError(f"Key Column '{key_col}'이 파일 B에 없습니다.")

    def _compare_dataframes(
        self,
        df_a: pd.DataFrame,
        df_b: pd.DataFrame,
    ) -> None:
        """DataFrame 비교 수행 (Vectorized)"""

        # 1. Key 컬럼 생성 (Tuple 기반 - 충돌 방지)
        # [Sprint 2.1] 기존 '|' 구분자 대신 Tuple 사용으로 특수문자 충돌 방지
        if not self._key_columns:
            raise ValueError("Key Columns must be defined")

        def create_key_series(df: pd.DataFrame) -> pd.Series:
            """Tuple 기반 안전한 Key 생성"""
            if len(self._key_columns) == 1:
                # 단일 컬럼: 문자열로 변환
                return df[self._key_columns[0]].astype(str)
            else:
                # 복합 컬럼: Tuple로 변환 (충돌 불가)
                return df[self._key_columns].apply(tuple, axis=1)

        df_a["_key"] = create_key_series(df_a)
        df_b["_key"] = create_key_series(df_b)

        # 2. Index 설정
        df_a_indexed = df_a.set_index("_key")
        df_b_indexed = df_b.set_index("_key")

        # 중복 Key 체크
        if df_a_indexed.index.duplicated().any():
            self._result.warnings.append(
                "파일 A에 중복된 Key가 존재합니다. 첫 번째 항목만 비교됩니다."
            )
            df_a_indexed = df_a_indexed.loc[~df_a_indexed.index.duplicated(keep="first")]

        if df_b_indexed.index.duplicated().any():
            self._result.warnings.append(
                "파일 B에 중복된 Key가 존재합니다. 첫 번째 항목만 비교됩니다."
            )
            df_b_indexed = df_b_indexed.loc[~df_b_indexed.index.duplicated(keep="first")]

        keys_a = set(df_a_indexed.index)
        keys_b = set(df_b_indexed.index)

        # 3. 삭제/추가된 항목 식별
        deleted_keys = keys_a - keys_b
        added_keys = keys_b - keys_a
        common_keys = keys_a & keys_b

        # 삭제된 항목 기록
        for key in deleted_keys:
            self._result.add_change(
                ChangeRecord(
                    key=key,
                    change_type=ChangeType.DELETED,
                    location=f"Row (Key: {key})",
                )
            )

        # 추가된 항목 기록
        for key in added_keys:
            self._result.add_change(
                ChangeRecord(
                    key=key,
                    change_type=ChangeType.ADDED,
                    location=f"Row (Key: {key})",
                )
            )

        # 4. 공통 항목 비교 (Vectorized-ish)
        if not common_keys:
            return

        # 공통 Key만 추출하여 정렬
        common_keys_list = list(common_keys)
        # DataFrame 정렬 (순서 일치 보장)
        df_a_common = df_a_indexed.loc[common_keys_list]
        df_b_common = df_b_indexed.loc[common_keys_list]

        # 비교할 컬럼 추출
        all_cols = set(df_a.columns) & set(df_b.columns)
        compare_cols = list(all_cols - set(self._key_columns) - self._ignore_columns - {"_key"})

        for col in compare_cols:
            # Series 추출
            col_a = df_a_common[col]
            col_b = df_b_common[col]

            # 1. 값이 같은지 먼저 체크 (Numeric, String 모두 포함)
            # == 연산은 NaN != NaN 이므로, fillna 처리가 필요하거나 equals 사용
            # 하지만 Numeric Tolerance 때문에 단순 == 로는 부족함.

            # Numeric 변환 시도
            try:
                num_a = pd.to_numeric(col_a, errors="coerce")
                num_b = pd.to_numeric(col_b, errors="coerce")

                # 둘 다 숫자인 경우: 차이가 tolerance보다 큰 경우만 True
                is_num = (~num_a.isna()) & (~num_b.isna())
                diff = (num_a - num_b).abs()
                changed_mask = is_num & (diff > self._numeric_tolerance)

                # 하나만 숫자인 경우 (Type Mismatch) or 둘 다 NaN이 아닌데 다른 경우
                # 문자열 비교를 위해 원본 데이터 사용
                # 간단하게: 값이 "다르다"고 판단되는 인덱스만 추출 후 반복문으로 정밀 검사
                # 완전 벡터화는 복잡하므로, "잠재적 변경"만 필터링하고 루프 도는 하이브리드 방식 적용

                # 단순 비교 (문자열 포함)
                # NaN == NaN 처리를 위해, 둘 다 NaN이면 False(변경없음) 처리

                # 객체 비교 (Object types)
                mask_diff = col_a.ne(col_b)

                # NaN 끼리는 변경 아님
                mask_nan = col_a.isna() & col_b.isna()
                mask_diff = mask_diff & (~mask_nan)

                if not mask_diff.any():
                    continue

                # 변경된 후보군에 대해서만 정밀 비교 (Format Ignore 등 적용)
                diff_indices = mask_diff[mask_diff].index

                for key in diff_indices:
                    val_a = col_a.at[key]
                    val_b = col_b.at[key]

                    if not self._values_equal(val_a, val_b):
                        self._result.add_change(
                            ChangeRecord(
                                key=key,
                                change_type=ChangeType.MODIFIED,
                                field_name=col,
                                old_value=val_a,
                                new_value=val_b,
                                location=f"Row (Key: {key}), Column: {col}",
                            )
                        )

            except Exception as e:
                logger.warning(f"Column '{col}' 비교 중 벡터화 최적화 실패, Fallback: {e}")
                # Fallback to loop for this column if something goes wrong
                for key in common_keys:
                    val_a = df_a_common.at[key, col]
                    val_b = df_b_common.at[key, col]
                    if not self._values_equal(val_a, val_b):
                        self._result.add_change(
                            ChangeRecord(
                                key=key,
                                change_type=ChangeType.MODIFIED,
                                field_name=col,
                                old_value=val_a,
                                new_value=val_b,
                                location=f"Row (Key: {key}), Column: {col}",
                            )
                        )

    def _values_equal(self, val_a: Any, val_b: Any) -> bool:
        """값 비교 (Tolerance 적용)"""
        # NaN 처리
        if pd.isna(val_a) and pd.isna(val_b):
            return True
        if pd.isna(val_a) or pd.isna(val_b):
            return False

        # 숫자 비교 (Tolerance 적용)
        try:
            num_a = float(val_a)
            num_b = float(val_b)
            return abs(num_a - num_b) <= self._numeric_tolerance
        except (ValueError, TypeError):
            pass

        # 문자열 비교 (Format Ignore: 쉼표, 공백 제거 후 비교)
        str_a = str(val_a).replace(",", "").replace(" ", "").strip()
        str_b = str(val_b).replace(",", "").replace(" ", "").strip()

        return str_a == str_b

    def export_report(
        self,
        output_path: Union[str, Path],
        format: str = "excel",
    ) -> Path:
        """비교 결과를 리포트로 내보냅니다."""
        if not self._result:
            raise ValueError("비교 결과가 없습니다. 먼저 compare()를 실행하세요.")

        output_path = Path(output_path)
        output_path.parent.mkdir(parents=True, exist_ok=True)

        if format == "excel":
            return self._export_excel_report(output_path)
        elif format == "json":
            return self._export_json_report(output_path)
        else:
            raise ValueError(f"지원하지 않는 형식: {format}")

    def _export_excel_report(self, output_path: Path) -> Path:
        """Excel 형식으로 리포트 내보내기"""
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

        wb = openpyxl.Workbook()

        # 요약 시트
        ws_summary = wb.active
        ws_summary.title = "요약"
        ws_summary["A1"] = "Excel 비교 결과 리포트"
        ws_summary["A1"].font = Font(bold=True, size=14)

        ws_summary["A3"] = "비교 대상 A"
        ws_summary["B3"] = self._result.source_a
        ws_summary["A4"] = "비교 대상 B"
        ws_summary["B4"] = self._result.source_b
        ws_summary["A5"] = "비교 일시"
        ws_summary["B5"] = self._result.compared_at.strftime("%Y-%m-%d %H:%M:%S")

        ws_summary["A7"] = "추가된 항목"
        ws_summary["B7"] = self._result.added_count
        ws_summary["A8"] = "삭제된 항목"
        ws_summary["B8"] = self._result.deleted_count
        ws_summary["A9"] = "수정된 항목"
        ws_summary["B9"] = self._result.modified_count

        # 상세 변경 내역 시트
        ws_changes = wb.create_sheet("변경 내역")
        headers = ["Key", "변경 유형", "필드명", "이전 값", "새 값", "위치"]

        # 스타일
        header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        added_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        deleted_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        modified_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")

        for col, header in enumerate(headers, 1):
            cell = ws_changes.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font

        for row_idx, change in enumerate(self._result.changes, 2):
            ws_changes.cell(row=row_idx, column=1, value=change.key)
            ws_changes.cell(row=row_idx, column=2, value=change.change_type.value)
            ws_changes.cell(row=row_idx, column=3, value=change.field_name or "-")
            ws_changes.cell(
                row=row_idx, column=4, value=str(change.old_value) if change.old_value else "-"
            )
            ws_changes.cell(
                row=row_idx, column=5, value=str(change.new_value) if change.new_value else "-"
            )
            ws_changes.cell(row=row_idx, column=6, value=change.location or "-")

            # 변경 유형별 색상
            fill = None
            if change.change_type == ChangeType.ADDED:
                fill = added_fill
            elif change.change_type == ChangeType.DELETED:
                fill = deleted_fill
            elif change.change_type == ChangeType.MODIFIED:
                fill = modified_fill

            if fill:
                for col in range(1, 7):
                    ws_changes.cell(row=row_idx, column=col).fill = fill

        # 열 너비 조정
        ws_changes.column_dimensions["A"].width = 25
        ws_changes.column_dimensions["B"].width = 12
        ws_changes.column_dimensions["C"].width = 20
        ws_changes.column_dimensions["D"].width = 25
        ws_changes.column_dimensions["E"].width = 25
        ws_changes.column_dimensions["F"].width = 35

        wb.save(output_path)
        logger.info(f"Excel 리포트 생성: {output_path}")

        return output_path

    def _export_json_report(self, output_path: Path) -> Path:
        """JSON 형식으로 리포트 내보내기"""
        import json

        with open(output_path, "w", encoding="utf-8") as f:
            json.dump(self._result.to_dict(), f, ensure_ascii=False, indent=2)

        logger.info(f"JSON 리포트 생성: {output_path}")
        return output_path
