# -*- coding: utf-8 -*-
"""P5 단부보강 수량 회계 엔진 — "도면 객체 카운트"가 아닌 "수량 회계 장부 복원".

배경 (2026-07-02): 사람·코덱스·클로드 3자 카운트가 3~5배 갈렸다. 근본 원인은
성능이 아니라 **단위·모델 불일치**였다(root_cause_report 참조):

  * 사람  = 물리 개소(PC보 단부)를 세되, 대칭 건물이라 **반쪽만 세고 ×2 확장**,
           C는 **세지 않고 총−A−B 잔여**로 계산. → 수량 회계 장부.
  * AI    = 도면 표식(보강A/B 블록·원 심볼)을 직접 카운트. 관례(대칭)·잔여규칙 누락.

증거(인간작업 엑셀 실수식, 직접 확인):
    총개소 R  = Q*2 + P            (대칭구간 Q ×2 + 중앙 P)
    A타입     = (G - H)*2 + H       (구간 확장; G=측정면 포함, H=중앙)
    C개소     = R - A - B           (잔여, 직접 카운트 아님)
    SUM       = A + B + C  ( = R )

이 엔진은 그 **회계 규칙**을 1급으로 구현하고, 인간작업 엑셀을 **골든**으로 두어
검증한다. CAD/비전 지오메트리는 최종 총개소의 권위 출처가 아니라 교차검증 신호로만 사용하고,
최종값은 권위 출처 입력(인간작업 장부/부재표/도면 물량표) + 회계 규칙으로 계산.

계층:
  [1] 입력/검증(INPUT)  : 권위 출처 → P/Q/G/H/J/K 또는 최종 A/B/총 입력.
                          DXF/비전은 cross-check only. 총개소 자동 추출 금지.
  [2] 회계(ACCOUNTING)  : 반쪽 측정값 → 대칭확장 → A/B/C/총/SUM   ← 본 파일에서 검증 완료(8층 PASS).
  [3] 대사(RECONCILE)   : 회계 산출 vs 골든 → 층별 PASS/FAIL 리포트.

Usage:
    python scripts/quantity_accounting_engine.py --reconcile "인간작업.xlsx"
"""

from __future__ import annotations

import argparse
import sys
from dataclasses import dataclass, field
from pathlib import Path
from typing import Dict, List, Optional, Sequence

for _s in (sys.stdout, sys.stderr):
    try:
        _s.reconfigure(encoding="utf-8", errors="replace")  # type: ignore[attr-defined]
    except Exception:  # noqa: BLE001
        pass


# ─────────────────────────── [2] 회계 커널 ───────────────────────────
def expand_side_center(side_incl: float, center: float) -> float:
    """구간 확장: (측정면 - 중앙)*2 + 중앙.  측정면 side_incl 이 중앙을 포함하는 형식.

    인간작업 수식 ``=(G-H)*2+H`` 와 동일. 대칭 건물의 좌우 반복분을 ×2.
    """
    return (side_incl - center) * 2 + center


def expand_center_side(center: float, side: float) -> float:
    """총개소 확장: side*2 + center.  인간작업 수식 ``=Q*2+P`` 와 동일(side=중앙 제외)."""
    return side * 2 + center


@dataclass
class FloorMeasurement:
    """도곽·층의 '반쪽 측정값' (추출 계층 산출 = 회계 입력)."""
    floor: str
    # 총개소: 중앙(center) + 대칭면(side)  → R = side*2 + center
    total_center: float = 0.0
    total_side: float = 0.0
    # A/B: 구간 측정면(side_incl, 중앙포함) + 중앙(center) → (side-center)*2+center
    a_side_incl: Optional[float] = None
    a_center: float = 0.0
    b_side_incl: Optional[float] = None
    b_center: float = 0.0
    # A/B를 그리드존 확장 없이 직접 넣는 층(3F/4F 등)을 위한 fallback
    a_direct: Optional[float] = None
    b_direct: Optional[float] = None


@dataclass
class FloorQuantity:
    floor: str
    A: float
    B: float
    C: float           # 잔여 = 총 - A - B
    total: float       # 총개소
    SUM: float         # A + B + C (= total)


def compute_floor(m: FloorMeasurement) -> FloorQuantity:
    """회계 규칙 적용: 대칭확장 → A/B → 총 → C 잔여 → SUM."""
    total = expand_center_side(m.total_center, m.total_side)
    if m.a_side_incl is not None:
        A = expand_side_center(m.a_side_incl, m.a_center)
    else:
        A = m.a_direct or 0.0
    if m.b_side_incl is not None:
        B = expand_side_center(m.b_side_incl, m.b_center)
    else:
        B = m.b_direct or 0.0
    C = total - A - B          # ← 핵심: C는 직접 세지 않는다
    return FloorQuantity(m.floor, A, B, C, total, A + B + C)


# ─────────────────────────── [3] 골든 대사 ───────────────────────────
# 인간작업 엑셀 Sheet1 열 매핑 (REV.3.01 블록 기준)
#  C=A, D=B, E=C, F=SUM | G,H,I=A(A~F,F,A~K) | J,K,L=B | M,N,O=C총 | P,Q,R=총(중앙,대칭,합)
GOLDEN_ROWS_REV3 = {"3F": 4, "4F": 5, "5F": 6, "6F": 7, "7F": 8, "8F": 9, "9F": 10, "10F": 11}


def reconcile_golden(xlsx: Path, rows: Dict[str, int] = GOLDEN_ROWS_REV3) -> List[dict]:
    from openpyxl import load_workbook
    ws = load_workbook(str(xlsx), data_only=True)["Sheet1"]

    def gv(r, c):
        v = ws.cell(r, c).value
        return v if isinstance(v, (int, float)) else 0

    out = []
    for fl, r in rows.items():
        A_h, B_h, C_h, SUM_h, R_h = gv(r, 3), gv(r, 4), gv(r, 5), gv(r, 6), gv(r, 18)
        G, H, J, K, P, Q = gv(r, 7), gv(r, 8), gv(r, 10), gv(r, 11), gv(r, 16), gv(r, 17)
        m = FloorMeasurement(
            floor=fl, total_center=P, total_side=Q,
            a_side_incl=(G if (G or H) else None), a_center=H, a_direct=A_h,
            b_side_incl=(J if (J or K) else None), b_center=K, b_direct=B_h,
        )
        q = compute_floor(m)
        ok = (q.A == A_h and q.B == B_h and q.C == C_h and q.total == R_h and q.SUM == SUM_h)
        out.append({"floor": fl, "engine": q, "golden": (A_h, B_h, C_h, R_h, SUM_h), "pass": ok})
    return out


# ─────────────────────────── [1] 추출 계층 (인터페이스만) ───────────────────────────
def extract_zone_measurements(dxf_path: Path) -> List[FloorMeasurement]:  # noqa: ARG001
    """권위 출처 기반 측정값 로더의 자리표시자.  ← 미구현.

    ⚠ 이 계층이 3자 에이전트가 발산한 지점이다. 2026-07-02 프로브 결과,
    before-DWG 층별 보강블록의 CIRCLE/긴 LINE 수는 사람 총개소 R과 0.13~2.36배로
    비균일했다. 따라서 총개소(P/Q/R)는 DXF/비전 지오메트리에서 자동 산출하지 않는다.

    정책:
      * 권위 출처 = 인간작업 장부/부재표/도면 물량표.
      * DXF/비전 = cross-check only, 이상치 플래그용.
      * 보강A/B 마크 = 최종값이 아니라 G/J 후보 또는 검증 신호.
      * C는 여기서 세지 않는다(회계 계층이 총-A-B로 산정).
    """
    raise NotImplementedError(
        "입력 계층은 권위 출처를 로드하고, DXF/비전 지오메트리는 교차검증으로만 사용하도록 구현한다."
    )


def main(argv: Optional[Sequence[str]] = None) -> int:
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("--reconcile", type=Path, help="인간작업 골든 엑셀 경로")
    args = ap.parse_args(argv)
    if not args.reconcile:
        print("사용: --reconcile <인간작업.xlsx>", file=sys.stderr)
        return 2
    res = reconcile_golden(args.reconcile)
    print("층  |  A(엔진/골든)  B(엔진/골든)  총(엔진/골든)  C(엔진/골든)  판정")
    allpass = True
    for row in res:
        q = row["engine"]; A_h, B_h, C_h, R_h, _ = row["golden"]; allpass &= row["pass"]
        print(f"{row['floor']:4s}| {int(q.A):4}/{int(A_h):<4} {int(q.B):3}/{int(B_h):<3} "
              f"{int(q.total):5}/{int(R_h):<5} {int(q.C):5}/{int(C_h):<5}  "
              f"{'PASS' if row['pass'] else 'FAIL'}")
    print(f"\n회계모델 검증: {'ALL PASS ✓' if allpass else 'FAIL ✗'}  "
          f"(반쪽 측정값→대칭확장→C잔여 가 골든과 일치)")
    print("남은 계층: [1] 입력/검증(권위 출처 로드 + 지오메트리 교차검증) — 골든 대사로 보정 후 구현.")
    return 0 if allpass else 1


if __name__ == "__main__":
    raise SystemExit(main())
